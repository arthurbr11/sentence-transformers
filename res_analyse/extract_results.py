#!/usr/bin/env python3
"""
Clean Results Extractor for Sentence Transformers
================================================

Extracts NDCG, FLOPS, and Active Dims metrics with:
- Sub-columns by context length (256, 512)
- Bold separators between custom and pretrained models
- Visual highlighting for best performers
- Timestamp-based file naming in timestamped subfolders
- Excel and HTML outputs
"""

from __future__ import annotations

import json
import os
import re
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Configuration
RESULTS_DIR = "/home/user/Desktop/project/sentence-transformers/results"
AGGREGATE_BASE_DIR = "/home/user/Desktop/project/sentence-transformers/res_analyse/aggregate"
CONTEXT_LENGTHS = [256, 512]
CUSTOM_MODEL_PREFIX = "models__"

DATASETS = [
    "NanoMSMARCO",
    "NanoNQ",
    "NanoFEVER",
    "NanoHotpotQA",
    "NanoFiQA2018",
    "NanoArguAna",
    "NanoSciFact",
    "NanoQuoraRetrieval",
    "NanoDBPedia",
    "NanoSCIDOCS",
    "NanoNFCorpus",
    "NanoClimateFEVER",
    "NanoTouche2020",
]

METRICS = {
    "ndcg": "dot-NDCG@10",
    "query_dims": "query_active_dims",
    "corpus_dims": "corpus_active_dims",
    "flops": "avg_flops",
}

# MTEB Configuration
MTEB_ENGLISH_DATASETS = [
    "ArguAna",
    "CQADupstackGamingRetrieval",
    "CQADupstackUnixRetrieval",
    "ClimateFEVERHardNegatives",
    "FEVERHardNegatives",
    "FiQA2018",
    "HotpotQAHardNegatives",
    "SCIDOCS",
    "TRECCOVID",
    "Touche2020Retrieval.v3",
]

# Language configurations for multilingual MTEB
LANGUAGE_CONFIGS = {
    "fra": {
        "tasks": [
            "BelebeleRetrieval",
            "MIRACLRetrievalHardNegatives",
            "StatcanDialogueDatasetRetrieval",
        ],
        "average_filename": "Average_FRA_RETRIEVAL_TASKS.json",
    },
    "spa": {
        "tasks": [
            "BelebeleRetrieval",
            "MIRACLRetrievalHardNegatives",
            "MLQARetrieval",
        ],
        "average_filename": "Average_SPA_RETRIEVAL_TASKS.json",
    },
    "deu": {
        "tasks": [
            "BelebeleRetrieval",
            "MIRACLRetrievalHardNegatives",
            "MLQARetrieval",
            "WikipediaRetrievalMultilingual",
        ],
        "average_filename": "Average_DEU_RETRIEVAL_TASKS.json",
    },
}

# Styling
BEST_FONT = Font(bold=True, color="FF0000")
TOP_10_FONT = Font(underline="single")
HEADER_FONT = Font(bold=True, size=12)
SEPARATOR_FILL = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
CONTEXT_FILL = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")


def clean_model_name(dir_name):
    """Extract clean model name and determine if custom"""
    is_custom = dir_name.startswith(CUSTOM_MODEL_PREFIX)
    name = re.sub(r"^(models__|naver__|prithivida__|opensearch-project__|ibm-granite__)", "", dir_name)

    # Extract checkpoint number if present and add it in readable format
    checkpoint_match = re.search(r"__checkpoint-(\d+)$", name)
    if checkpoint_match:
        checkpoint_num = checkpoint_match.group(1)
        # Remove the checkpoint from the base name and add it in readable format
        base_name = re.sub(r"__checkpoint-\d+$", "", name)
        name = f"{base_name} (ckpt {checkpoint_num})"

    return name, is_custom


def load_results(model_dir, context_length):
    """Load mean and dataset results for a model"""
    results_path = os.path.join(RESULTS_DIR, model_dir, f"NanoBEIR_{context_length}")

    # Load mean results
    mean_file = os.path.join(results_path, "NanoBEIR_evaluation_mean_results.csv")
    mean_data = None
    if os.path.exists(mean_file):
        mean_data = load_csv_with_header_body_mismatch(mean_file)

    # Load dataset results
    dataset_data = {}
    for dataset in DATASETS:
        file_path = os.path.join(results_path, f"Information-Retrieval_evaluation_{dataset}_results.csv")
        if os.path.exists(file_path):
            dataset_data[dataset] = load_csv_with_header_body_mismatch(file_path)

    return mean_data, dataset_data


def load_csv_with_header_body_mismatch(file_path):
    """Load CSV handling cases where header has more columns than body data"""
    try:
        # Read the file manually to handle header/body mismatch
        with open(file_path) as f:
            lines = f.readlines()

        if len(lines) < 2:
            return None

        # Parse header and data
        header = lines[0].strip().split(",")
        data_line = lines[-1].strip().split(",")  # Take last line (most recent results)

        # If header and data lengths match, use normal loading
        if len(header) == len(data_line):
            df = pd.read_csv(file_path)
            return df.iloc[-1]

        # Handle mismatch: first 2 columns are good, then align from the end
        if len(data_line) < len(header) and len(data_line) >= 2:
            # Keep first 2 columns as-is
            aligned_header = header[:2]
            aligned_data = data_line[:2]

            # For the remaining data, align from the end
            remaining_data_count = len(data_line) - 2
            if remaining_data_count > 0:
                # Take the last N columns from header to match remaining data
                aligned_header.extend(header[-remaining_data_count:])
                aligned_data.extend(data_line[2:])

            # Create a Series with aligned data
            try:
                # Convert data to appropriate types
                converted_data = []
                for i, value in enumerate(aligned_data):
                    try:
                        # Try to convert to float
                        converted_data.append(float(value))
                    except ValueError:
                        # Keep as string if conversion fails
                        converted_data.append(value)

                return pd.Series(converted_data, index=aligned_header)
            except Exception:
                return None

        # Fallback: try normal loading anyway
        try:
            df = pd.read_csv(file_path)
            return df.iloc[-1]
        except Exception:
            return None

    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return None


def load_mteb_english_results(model_dir, context_length):
    """Load MTEB English results for a model"""
    mteb_path = os.path.join(RESULTS_DIR, model_dir, f"MTEB_{context_length}")

    if not os.path.exists(mteb_path):
        return None, {}

    # Load English datasets
    english_datasets = {}

    for dataset in MTEB_ENGLISH_DATASETS:
        json_file = os.path.join(mteb_path, f"{dataset}.json")
        if os.path.exists(json_file):
            try:
                with open(json_file) as f:
                    data = json.load(f)
                    if "scores" in data and "test" in data["scores"] and data["scores"]["test"]:
                        english_datasets[dataset] = {
                            "main_score": data["scores"]["test"][0]["main_score"],
                            "ndcg_at_10": data["scores"]["test"][0].get(
                                "ndcg_at_10", data["scores"]["test"][0]["main_score"]
                            ),
                        }
            except (json.JSONDecodeError, KeyError, IndexError):
                continue

    # Load English average
    english_average = None
    eng_avg_file = os.path.join(mteb_path, "Average_EN_V2_RETRIEVAL_TASKS.json")
    if os.path.exists(eng_avg_file):
        try:
            with open(eng_avg_file) as f:
                data = json.load(f)
                if "scores" in data and "test" in data["scores"] and data["scores"]["test"]:
                    english_average = {
                        "main_score": data["scores"]["test"][0]["main_score"],
                        "ndcg_at_10": data["scores"]["test"][0].get(
                            "ndcg_at_10", data["scores"]["test"][0]["main_score"]
                        ),
                    }
        except (json.JSONDecodeError, KeyError, IndexError):
            pass

    # If no official average, calculate from individual datasets
    if english_average is None and english_datasets:
        scores = [data["main_score"] for data in english_datasets.values() if "main_score" in data]
        if scores:
            english_average = {"main_score": sum(scores) / len(scores), "ndcg_at_10": sum(scores) / len(scores)}

    return english_average, english_datasets


def load_mteb_multilingual_results(model_dir, context_length):
    """Load MTEB multilingual results for a model"""
    mteb_path = os.path.join(RESULTS_DIR, model_dir, f"MTEB_{context_length}")

    if not os.path.exists(mteb_path):
        return {}, {}, None

    # Load language-specific averages and datasets
    language_averages = {}
    language_datasets = {}

    for lang_code, config in LANGUAGE_CONFIGS.items():
        # Load language average
        avg_file = os.path.join(mteb_path, config["average_filename"])
        if os.path.exists(avg_file):
            try:
                with open(avg_file) as f:
                    data = json.load(f)
                    if "scores" in data and "test" in data["scores"] and data["scores"]["test"]:
                        language_averages[lang_code] = {
                            "main_score": data["scores"]["test"][0]["main_score"],
                            "ndcg_at_10": data["scores"]["test"][0].get(
                                "ndcg_at_10", data["scores"]["test"][0]["main_score"]
                            ),
                        }
            except (json.JSONDecodeError, KeyError, IndexError):
                continue

        # Load individual datasets for this language
        lang_dataset_data = {}
        for dataset in config["tasks"]:
            json_file = os.path.join(mteb_path, f"{dataset}.json")
            if os.path.exists(json_file):
                try:
                    with open(json_file) as f:
                        data = json.load(f)
                        if "scores" in data and "test" in data["scores"] and data["scores"]["test"]:
                            # For multilingual datasets, find the score for this language
                            for result in data["scores"]["test"]:
                                # Match language code or language name
                                result_langs = result.get("languages", [])
                                if any(lang_code in lang.lower() for lang in result_langs):
                                    lang_dataset_data[dataset] = {
                                        "main_score": result["main_score"],
                                        "ndcg_at_10": result.get("ndcg_at_10", result["main_score"]),
                                    }
                                    break
                except (json.JSONDecodeError, KeyError, IndexError):
                    continue

        if lang_dataset_data:
            language_datasets[lang_code] = lang_dataset_data

    # Calculate global average across all languages
    global_average = None
    if language_averages:
        all_scores = [data["main_score"] for data in language_averages.values()]
        if all_scores:
            global_average = {
                "main_score": sum(all_scores) / len(all_scores),
                "ndcg_at_10": sum(all_scores) / len(all_scores),
            }

    return language_averages, language_datasets, global_average


def format_value(value, decimal_places=2):
    """Format numeric value"""
    if pd.isna(value):
        return "N/A"
    return f"{value:.{decimal_places}f}"


def extract_all_data():
    """Extract all model data"""
    models = {"custom": [], "pretrained": []}
    all_data = {}

    for model_dir in sorted(os.listdir(RESULTS_DIR)):
        if not os.path.isdir(os.path.join(RESULTS_DIR, model_dir)):
            continue

        model_name, is_custom = clean_model_name(model_dir)
        category = "custom" if is_custom else "pretrained"
        models[category].append(model_name)

        all_data[model_name] = {"type": category}

        for context_length in CONTEXT_LENGTHS:
            # Load NanoBEIR results
            mean_data, dataset_data = load_results(model_dir, context_length)

            # Load MTEB results only for 256 context
            if context_length == 256:
                # Load MTEB English results
                english_average, english_datasets = load_mteb_english_results(model_dir, context_length)

                # Load MTEB multilingual results
                multilingual_averages, multilingual_datasets, global_average = load_mteb_multilingual_results(
                    model_dir, context_length
                )
            else:
                # No MTEB data for 512 context
                english_average, english_datasets = None, {}
                multilingual_averages, multilingual_datasets, global_average = {}, {}, None

            all_data[model_name][context_length] = {
                "mean": mean_data,
                "datasets": dataset_data,
                "mteb_english_average": english_average,
                "mteb_english_datasets": english_datasets,
                "mteb_multilingual_averages": multilingual_averages,
                "mteb_multilingual_datasets": multilingual_datasets,
                "mteb_global_average": global_average,
            }

    return models, all_data


def create_summary_table(models, all_data):
    """Create summary table with sub-columns"""
    rows = []

    # Add custom models
    for model_name in sorted(models["custom"]):
        row = {"Model": model_name, "Type": "Custom"}

        for context in CONTEXT_LENGTHS:
            mean_data = all_data[model_name][context]["mean"]
            mteb_english_average = all_data[model_name][context]["mteb_english_average"]

            if mean_data is not None:
                row[f"{context}_NDCG"] = format_value(mean_data.get(METRICS["ndcg"], np.nan) * 100)
                row[f"{context}_FLOPS"] = format_value(mean_data.get(METRICS["flops"], np.nan))
                row[f"{context}_QueryDims"] = format_value(mean_data.get(METRICS["query_dims"], np.nan), 1)
                row[f"{context}_CorpusDims"] = format_value(mean_data.get(METRICS["corpus_dims"], np.nan), 1)
            else:
                for metric in ["NDCG", "FLOPS", "QueryDims", "CorpusDims"]:
                    row[f"{context}_{metric}"] = "N/A"

            # Add MTEB English average
            if mteb_english_average is not None:
                row[f"{context}_MTEB"] = format_value(mteb_english_average.get("main_score", np.nan) * 100)
            else:
                row[f"{context}_MTEB"] = "N/A"
        rows.append(row)

    # Add separator
    if models["custom"] and models["pretrained"]:
        sep_row = {"Model": "--- PRETRAINED MODELS ---", "Type": "Separator"}
        for context in CONTEXT_LENGTHS:
            for metric in ["NDCG", "FLOPS", "QueryDims", "CorpusDims", "MTEB"]:
                sep_row[f"{context}_{metric}"] = "---"
        rows.append(sep_row)

    # Add pretrained models
    for model_name in sorted(models["pretrained"]):
        row = {"Model": model_name, "Type": "Pretrained"}

        for context in CONTEXT_LENGTHS:
            mean_data = all_data[model_name][context]["mean"]
            mteb_english_average = all_data[model_name][context]["mteb_english_average"]

            if mean_data is not None:
                row[f"{context}_NDCG"] = format_value(mean_data.get(METRICS["ndcg"], np.nan) * 100)
                row[f"{context}_FLOPS"] = format_value(mean_data.get(METRICS["flops"], np.nan))
                row[f"{context}_QueryDims"] = format_value(mean_data.get(METRICS["query_dims"], np.nan), 1)
                row[f"{context}_CorpusDims"] = format_value(mean_data.get(METRICS["corpus_dims"], np.nan), 1)
            else:
                for metric in ["NDCG", "FLOPS", "QueryDims", "CorpusDims"]:
                    row[f"{context}_{metric}"] = "N/A"

            # Add MTEB English average
            if mteb_english_average is not None:
                row[f"{context}_MTEB"] = format_value(mteb_english_average.get("main_score", np.nan) * 100)
            else:
                row[f"{context}_MTEB"] = "N/A"
        rows.append(row)

    return pd.DataFrame(rows)


def create_dataset_table(models, all_data, context_length):
    """Create dataset table for specific context length"""
    rows = []

    # Custom models
    for model_name in sorted(models["custom"]):
        row = {"Model": model_name, "Type": "Custom"}

        mean_data = all_data[model_name][context_length]["mean"]
        dataset_data = all_data[model_name][context_length]["datasets"]

        # Add dataset columns
        for dataset in DATASETS:
            if dataset in dataset_data:
                data = dataset_data[dataset]
                ndcg = format_value(data.get(METRICS["ndcg"], np.nan) * 100)
                flops = format_value(data.get(METRICS["flops"], np.nan))
                row[dataset] = f"{ndcg} ({flops})"
            else:
                row[dataset] = "N/A"

        # Add average
        if mean_data is not None:
            ndcg = format_value(mean_data.get(METRICS["ndcg"], np.nan) * 100)
            flops = format_value(mean_data.get(METRICS["flops"], np.nan))
            row["Average"] = f"{ndcg} ({flops})"
        else:
            row["Average"] = "N/A"

        rows.append(row)

    # Separator
    if models["custom"] and models["pretrained"]:
        sep_row = {"Model": "--- PRETRAINED MODELS ---", "Type": "Separator"}
        for col in DATASETS + ["Average"]:
            sep_row[col] = "---"
        rows.append(sep_row)

    # Pretrained models
    for model_name in sorted(models["pretrained"]):
        row = {"Model": model_name, "Type": "Pretrained"}

        mean_data = all_data[model_name][context_length]["mean"]
        dataset_data = all_data[model_name][context_length]["datasets"]

        # Add dataset columns
        for dataset in DATASETS:
            if dataset in dataset_data:
                data = dataset_data[dataset]
                ndcg = format_value(data.get(METRICS["ndcg"], np.nan) * 100)
                flops = format_value(data.get(METRICS["flops"], np.nan))
                row[dataset] = f"{ndcg} ({flops})"
            else:
                row[dataset] = "N/A"

        # Add average
        if mean_data is not None:
            ndcg = format_value(mean_data.get(METRICS["ndcg"], np.nan) * 100)
            flops = format_value(mean_data.get(METRICS["flops"], np.nan))
            row["Average"] = f"{ndcg} ({flops})"
        else:
            row["Average"] = "N/A"

        rows.append(row)

    return pd.DataFrame(rows)


def create_mteb_english_table(models, all_data, context_length):
    """Create MTEB English table for specific context length with individual datasets"""
    rows = []

    # Custom models
    for model_name in sorted(models["custom"]):
        row = {"Model": model_name, "Type": "Custom"}

        mteb_english_average = all_data[model_name][context_length]["mteb_english_average"]
        mteb_english_datasets = all_data[model_name][context_length]["mteb_english_datasets"]

        # Add dataset columns
        for dataset in MTEB_ENGLISH_DATASETS:
            if dataset in mteb_english_datasets:
                score = format_value(mteb_english_datasets[dataset].get("main_score", np.nan) * 100)
                row[dataset] = score
            else:
                row[dataset] = "N/A"

        # Add English average
        if mteb_english_average is not None:
            row["English_Average"] = format_value(mteb_english_average.get("main_score", np.nan) * 100)
        else:
            row["English_Average"] = "N/A"

        rows.append(row)

    # Separator
    if models["custom"] and models["pretrained"]:
        sep_row = {"Model": "--- PRETRAINED MODELS ---", "Type": "Separator"}
        for col in MTEB_ENGLISH_DATASETS + ["English_Average"]:
            sep_row[col] = "---"
        rows.append(sep_row)

    # Pretrained models
    for model_name in sorted(models["pretrained"]):
        row = {"Model": model_name, "Type": "Pretrained"}

        mteb_english_average = all_data[model_name][context_length]["mteb_english_average"]
        mteb_english_datasets = all_data[model_name][context_length]["mteb_english_datasets"]

        # Add dataset columns
        for dataset in MTEB_ENGLISH_DATASETS:
            if dataset in mteb_english_datasets:
                score = format_value(mteb_english_datasets[dataset].get("main_score", np.nan) * 100)
                row[dataset] = score
            else:
                row[dataset] = "N/A"

        # Add English average
        if mteb_english_average is not None:
            row["English_Average"] = format_value(mteb_english_average.get("main_score", np.nan) * 100)
        else:
            row["English_Average"] = "N/A"

        rows.append(row)

    return pd.DataFrame(rows)


def create_mteb_multilingual_table(models, all_data, context_length):
    """Create MTEB multilingual table with per-language averages"""
    rows = []

    # Check if any models have multilingual data
    has_multilingual_data = False
    for model_name in models["custom"] + models["pretrained"]:
        multilingual_averages = all_data[model_name][context_length]["mteb_multilingual_averages"]
        if multilingual_averages:
            has_multilingual_data = True
            break

    if not has_multilingual_data:
        # No multilingual data available
        return pd.DataFrame()

    # Custom models
    for model_name in sorted(models["custom"]):
        row = {"Model": model_name, "Type": "Custom"}

        multilingual_averages = all_data[model_name][context_length]["mteb_multilingual_averages"]
        global_average = all_data[model_name][context_length]["mteb_global_average"]

        # Add language-specific averages
        for lang_code in sorted(LANGUAGE_CONFIGS.keys()):
            if lang_code in multilingual_averages:
                score = format_value(multilingual_averages[lang_code].get("main_score", np.nan) * 100)
                row[f"{lang_code.upper()}_Avg"] = score
            else:
                row[f"{lang_code.upper()}_Avg"] = "N/A"

        # Add global average
        if global_average is not None:
            row["Global_Average"] = format_value(global_average.get("main_score", np.nan) * 100)
        else:
            row["Global_Average"] = "N/A"

        rows.append(row)

    # Separator
    if models["custom"] and models["pretrained"]:
        sep_row = {"Model": "--- PRETRAINED MODELS ---", "Type": "Separator"}
        for lang_code in sorted(LANGUAGE_CONFIGS.keys()):
            sep_row[f"{lang_code.upper()}_Avg"] = "---"
        sep_row["Global_Average"] = "---"
        rows.append(sep_row)

    # Pretrained models
    for model_name in sorted(models["pretrained"]):
        row = {"Model": model_name, "Type": "Pretrained"}

        multilingual_averages = all_data[model_name][context_length]["mteb_multilingual_averages"]
        global_average = all_data[model_name][context_length]["mteb_global_average"]

        # Add language-specific averages
        for lang_code in sorted(LANGUAGE_CONFIGS.keys()):
            if lang_code in multilingual_averages:
                score = format_value(multilingual_averages[lang_code].get("main_score", np.nan) * 100)
                row[f"{lang_code.upper()}_Avg"] = score
            else:
                row[f"{lang_code.upper()}_Avg"] = "N/A"

        # Add global average
        if global_average is not None:
            row["Global_Average"] = format_value(global_average.get("main_score", np.nan) * 100)
        else:
            row["Global_Average"] = "N/A"

        rows.append(row)

    return pd.DataFrame(rows)


def create_mteb_multilingual_datasets_table(models, all_data, context_length):
    """Create MTEB multilingual table with individual dataset scores per language"""
    rows = []

    # Check if any models have multilingual data
    has_multilingual_data = False
    for model_name in models["custom"] + models["pretrained"]:
        multilingual_datasets = all_data[model_name][context_length]["mteb_multilingual_datasets"]
        if multilingual_datasets:
            has_multilingual_data = True
            break

    if not has_multilingual_data:
        return pd.DataFrame()

    # Collect all possible datasets across all languages
    all_datasets = set()
    for lang_code, config in LANGUAGE_CONFIGS.items():
        all_datasets.update(config["tasks"])
    all_datasets = sorted(list(all_datasets))

    # Custom models
    for model_name in sorted(models["custom"]):
        row = {"Model": model_name, "Type": "Custom"}

        multilingual_datasets = all_data[model_name][context_length]["mteb_multilingual_datasets"]
        multilingual_averages = all_data[model_name][context_length]["mteb_multilingual_averages"]
        global_average = all_data[model_name][context_length]["mteb_global_average"]

        # Add columns for each language and dataset combination
        for lang_code in sorted(LANGUAGE_CONFIGS.keys()):
            # Add language average
            if lang_code in multilingual_averages:
                score = format_value(multilingual_averages[lang_code].get("main_score", np.nan) * 100)
                row[f"{lang_code.upper()}_Avg"] = score
            else:
                row[f"{lang_code.upper()}_Avg"] = "N/A"

            # Add individual datasets for this language
            lang_datasets = multilingual_datasets.get(lang_code, {})
            for dataset in LANGUAGE_CONFIGS[lang_code]["tasks"]:
                if dataset in lang_datasets:
                    score = format_value(lang_datasets[dataset].get("main_score", np.nan) * 100)
                    row[f"{lang_code.upper()}_{dataset}"] = score
                else:
                    row[f"{lang_code.upper()}_{dataset}"] = "N/A"

        # Add global average
        if global_average is not None:
            row["Global_Average"] = format_value(global_average.get("main_score", np.nan) * 100)
        else:
            row["Global_Average"] = "N/A"

        rows.append(row)

    # Separator
    if models["custom"] and models["pretrained"]:
        sep_row = {"Model": "--- PRETRAINED MODELS ---", "Type": "Separator"}
        for lang_code in sorted(LANGUAGE_CONFIGS.keys()):
            sep_row[f"{lang_code.upper()}_Avg"] = "---"
            for dataset in LANGUAGE_CONFIGS[lang_code]["tasks"]:
                sep_row[f"{lang_code.upper()}_{dataset}"] = "---"
        sep_row["Global_Average"] = "---"
        rows.append(sep_row)

    # Pretrained models
    for model_name in sorted(models["pretrained"]):
        row = {"Model": model_name, "Type": "Pretrained"}

        multilingual_datasets = all_data[model_name][context_length]["mteb_multilingual_datasets"]
        multilingual_averages = all_data[model_name][context_length]["mteb_multilingual_averages"]
        global_average = all_data[model_name][context_length]["mteb_global_average"]

        # Add columns for each language and dataset combination
        for lang_code in sorted(LANGUAGE_CONFIGS.keys()):
            # Add language average
            if lang_code in multilingual_averages:
                score = format_value(multilingual_averages[lang_code].get("main_score", np.nan) * 100)
                row[f"{lang_code.upper()}_Avg"] = score
            else:
                row[f"{lang_code.upper()}_Avg"] = "N/A"

            # Add individual datasets for this language
            lang_datasets = multilingual_datasets.get(lang_code, {})
            for dataset in LANGUAGE_CONFIGS[lang_code]["tasks"]:
                if dataset in lang_datasets:
                    score = format_value(lang_datasets[dataset].get("main_score", np.nan) * 100)
                    row[f"{lang_code.upper()}_{dataset}"] = score
                else:
                    row[f"{lang_code.upper()}_{dataset}"] = "N/A"

        # Add global average
        if global_average is not None:
            row["Global_Average"] = format_value(global_average.get("main_score", np.nan) * 100)
        else:
            row["Global_Average"] = "N/A"

        rows.append(row)

    return pd.DataFrame(rows)


def identify_best_performers(df):
    """Find best and top 10% performers"""
    performance = {}

    for col in df.columns:
        if col in ["Model", "Type"]:
            continue

        valid_data = df[(df["Type"] != "Separator") & (df[col] != "---") & (df[col] != "N/A")]
        if len(valid_data) == 0:
            continue

        # Determine if higher or lower is better
        is_higher_better = (
            "NDCG" in col
            or "MTEB" in col
            or any(d in col for d in DATASETS)
            or any(d in col for d in MTEB_ENGLISH_DATASETS)
            or col == "Average"
            or "Avg" in col
            or "English_Average" in col
            or "Global_Average" in col
            or any(lang in col for lang in ["FRA", "SPA", "DEU"])
        )
        ascending = not is_higher_better  # NDCG/MTEB higher is better, FLOPS lower is better

        # Extract numeric values
        numeric_vals = []
        for idx, value in valid_data[col].items():
            try:
                if "(" in str(value):
                    num_val = float(str(value).split("(")[0].strip())
                else:
                    num_val = float(value)
                numeric_vals.append((idx, num_val))
            except (ValueError, TypeError):
                continue

        if not numeric_vals:
            continue

        # Sort and identify best/top 10%
        sorted_vals = sorted(numeric_vals, key=lambda x: x[1], reverse=not ascending)
        best_idx = sorted_vals[0][0]
        top_10_count = max(1, len(sorted_vals) // 10)
        top_10_indices = [idx for idx, _ in sorted_vals[:top_10_count]]

        performance[col] = {"best": best_idx, "top_10": top_10_indices}

    return performance


def save_excel_with_formatting(summary_df, dataset_dfs, perf_data, output_dir):
    """Save to Excel with proper formatting"""
    filename = "results.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")

    # Create multi-level headers
    ws_summary.cell(1, 1, "Model")
    ws_summary.cell(1, 2, "Type")
    ws_summary.merge_cells("A1:A2")
    ws_summary.merge_cells("B1:B2")

    col_idx = 3
    col_mapping = {}

    for context in CONTEXT_LENGTHS:
        start_col = col_idx
        ws_summary.cell(1, col_idx, f"Context {context}")

        for metric in ["NDCG", "FLOPS", "QueryDims", "CorpusDims"]:
            ws_summary.cell(2, col_idx, metric)
            col_mapping[f"{context}_{metric}"] = col_idx
            col_idx += 1

        end_col = col_idx - 1
        ws_summary.merge_cells(f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1")

    # Add data
    for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=3):
        ws_summary.cell(row_idx, 1, row["Model"])
        ws_summary.cell(row_idx, 2, row["Type"])

        for col_name, col_pos in col_mapping.items():
            value = row.get(col_name, "N/A")
            cell = ws_summary.cell(row_idx, col_pos, value)

            # Apply highlighting
            if col_name in perf_data.get("summary", {}) and row["Type"] != "Separator":
                col_perf = perf_data["summary"][col_name]
                df_row_idx = row_idx - 3

                if df_row_idx == col_perf.get("best"):
                    cell.font = BEST_FONT
                elif df_row_idx in col_perf.get("top_10", []):
                    cell.font = TOP_10_FONT

    # Format headers
    for col in range(1, col_idx):
        ws_summary.cell(1, col).font = HEADER_FONT
        ws_summary.cell(1, col).alignment = Alignment(horizontal="center")
        ws_summary.cell(2, col).font = HEADER_FONT
        ws_summary.cell(2, col).alignment = Alignment(horizontal="center")
        if col > 2:
            ws_summary.cell(1, col).fill = CONTEXT_FILL

    # Format separator rows
    for row_idx in range(3, len(summary_df) + 3):
        if summary_df.iloc[row_idx - 3]["Type"] == "Separator":
            for col in range(1, col_idx):
                cell = ws_summary.cell(row_idx, col)
                cell.fill = SEPARATOR_FILL
                cell.font = Font(bold=True)

    # Auto-adjust column widths
    for col in range(1, col_idx):
        ws_summary.column_dimensions[get_column_letter(col)].width = 15

    # Dataset sheets
    for context, df in dataset_dfs.items():
        ws = wb.create_sheet(f"Datasets_{context}")

        # Add headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(1, col_idx, col_name).font = HEADER_FONT

        # Add data
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row_idx, col_idx, value)

                # Apply highlighting for separator rows
                if row["Type"] == "Separator":
                    cell.fill = SEPARATOR_FILL
                    cell.font = Font(bold=True)

        # Auto-adjust widths
        for col in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    wb.save(filepath)
    print(f"Saved Excel: {filename}")


def save_html_with_tabs(
    summary_df,
    dataset_dfs,
    mteb_dfs,
    mteb_multilingual_dfs,
    mteb_multilingual_datasets_dfs,
    perf_data,
    output_dir,
    timestamp,
):
    """Save to HTML with tabbed interface"""
    filename = "results.html"
    filepath = os.path.join(output_dir, filename)

    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>Sentence Transformers Results - {timestamp}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 20px; }}
            .tabs {{ display: flex; background-color: #f1f1f1; border-radius: 8px 8px 0 0; }}
            .tab {{ background-color: inherit; border: none; outline: none; cursor: pointer;
                   padding: 14px 20px; transition: 0.3s; font-size: 16px; }}
            .tab:hover {{ background-color: #ddd; }}
            .tab.active {{ background-color: #007acc; color: white; }}
            .tabcontent {{ display: none; padding: 20px; border: 1px solid #ccc; border-top: none; }}
            .tabcontent.active {{ display: block; }}
            table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: center; }}
            th {{ background-color: #f2f2f2; font-weight: bold; }}
            .context-header {{ background-color: #e6f3ff; font-weight: bold; font-size: 14px; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
            .separator {{ background-color: #cccccc; font-weight: bold; }}
            .best {{ color: red; font-weight: bold; }}
            .top10 {{ text-decoration: underline; }}
            .model-name {{ text-align: left; font-weight: bold; }}
            h1 {{ color: #333; text-align: center; margin-bottom: 30px; }}
            h2 {{ color: #666; margin-bottom: 20px; }}
            .timestamp {{ text-align: center; color: #888; margin-bottom: 20px; }}
        </style>
        <script>
            function openTab(evt, tabName) {{
                var i, tabcontent, tabs;
                tabcontent = document.getElementsByClassName("tabcontent");
                for (i = 0; i < tabcontent.length; i++) {{
                    tabcontent[i].classList.remove("active");
                }}
                tabs = document.getElementsByClassName("tab");
                for (i = 0; i < tabs.length; i++) {{
                    tabs[i].classList.remove("active");
                }}
                document.getElementById(tabName).classList.add("active");
                evt.currentTarget.classList.add("active");
            }}
        </script>
    </head>
    <body>
        <h1>🚀 Sentence Transformers Model Results</h1>
        <div class="timestamp">Generated on: {datetime.now().strftime("%B %d, %Y at %H:%M:%S")}</div>

        <div class="tabs">
            <button class="tab active" onclick="openTab(event, 'summary')">📊 Summary</button>
            <button class="tab" onclick="openTab(event, 'datasets_256')">📋 NanoBEIR 256</button>
            <button class="tab" onclick="openTab(event, 'datasets_512')">📋 NanoBEIR 512</button>
            <button class="tab" onclick="openTab(event, 'mteb_256')">🎯 MTEB English 256</button>
            {'<button class="tab" onclick="openTab(event, \'mteb_multilingual_256\')">🌍 MTEB Multilingual Avg 256</button>' if not mteb_multilingual_dfs[256].empty else ''}
            {'<button class="tab" onclick="openTab(event, \'mteb_multilingual_datasets_256\')">📊 MTEB Multilingual Datasets 256</button>' if not mteb_multilingual_datasets_dfs[256].empty else ''}
        </div>

        <div id="summary" class="tabcontent active">
            <h2>Summary - Average Metrics by Context Length (including MTEB)</h2>
            {df_to_html_with_subcolumns(summary_df, perf_data.get("summary", {}))}
        </div>

        <div id="datasets_256" class="tabcontent">
            <h2>NanoBEIR Individual Datasets - Context Length 256</h2>
            {df_to_html_simple(dataset_dfs[256], perf_data.get("datasets_256", {}))}
        </div>

        <div id="datasets_512" class="tabcontent">
            <h2>NanoBEIR Individual Datasets - Context Length 512</h2>
            {df_to_html_simple(dataset_dfs[512], perf_data.get("datasets_512", {}))}
        </div>

        <div id="mteb_256" class="tabcontent">
            <h2>MTEB English Individual Datasets - Context Length 256</h2>
            {df_to_html_simple(mteb_dfs[256], perf_data.get("mteb_256", {}))}
        </div>

        {f'<div id="mteb_multilingual_256" class="tabcontent"><h2>MTEB Multilingual Language Averages - Context Length 256</h2>{df_to_html_simple(mteb_multilingual_dfs[256], perf_data.get("mteb_multilingual_256", {}))}</div>' if not mteb_multilingual_dfs[256].empty else ''}

        {f'<div id="mteb_multilingual_datasets_256" class="tabcontent"><h2>MTEB Multilingual Individual Datasets - Context Length 256</h2>{df_to_html_simple(mteb_multilingual_datasets_dfs[256], perf_data.get("mteb_multilingual_datasets_256", {}))}</div>' if not mteb_multilingual_datasets_dfs[256].empty else ''}

    </body>
    </html>
    """

    with open(filepath, "w") as f:
        f.write(html_content)
    print(f"Saved HTML: {filename}")


def df_to_html_with_subcolumns(df, perf_data):
    """Convert summary DataFrame to HTML with sub-column headers"""
    html = "<table>"

    # Create multi-level header
    html += "<tr>"
    html += "<th rowspan='2'>Model</th>"
    html += "<th rowspan='2'>Type</th>"

    # Context length headers
    for context in CONTEXT_LENGTHS:
        html += f"<th colspan='5' class='context-header'>Context {context}</th>"
    html += "</tr>"

    # Sub-headers
    html += "<tr>"
    for context in CONTEXT_LENGTHS:
        html += "<th>NDCG</th><th>FLOPS</th><th>QueryDims</th><th>CorpusDims</th><th>MTEB</th>"
    html += "</tr>"

    # Data rows
    for idx, row in df.iterrows():
        row_class = "separator" if row["Type"] == "Separator" else ""
        html += f"<tr class='{row_class}'>"

        html += f"<td class='model-name'>{row['Model']}</td>"
        html += f"<td>{row['Type']}</td>"

        # Add metric values
        for context in CONTEXT_LENGTHS:
            for metric in ["NDCG", "FLOPS", "QueryDims", "CorpusDims", "MTEB"]:
                col_name = f"{context}_{metric}"
                value = row.get(col_name, "N/A")

                cell_class = ""
                if col_name in perf_data and row["Type"] != "Separator":
                    col_perf = perf_data[col_name]
                    if idx == col_perf.get("best"):
                        cell_class = "best"
                    elif idx in col_perf.get("top_10", []):
                        cell_class = "top10"

                html += f"<td class='{cell_class}'>{value}</td>"

        html += "</tr>"

    html += "</table>"
    return html


def df_to_html_simple(df, perf_data):
    """Convert DataFrame to HTML with simple styling"""
    html = "<table>"

    # Header
    html += "<tr>"
    for col in df.columns:
        html += f"<th>{col}</th>"
    html += "</tr>"

    # Rows
    for idx, row in df.iterrows():
        row_class = "separator" if row["Type"] == "Separator" else ""
        html += f"<tr class='{row_class}'>"

        for col_name, value in row.items():
            cell_class = ""

            # Apply performance highlighting
            if col_name in perf_data and row["Type"] != "Separator":
                col_perf = perf_data[col_name]
                if idx == col_perf.get("best"):
                    cell_class = "best"
                elif idx in col_perf.get("top_10", []):
                    cell_class = "top10"

            # Special formatting for model names
            if col_name == "Model":
                cell_class += " model-name"

            html += f"<td class='{cell_class}'>{value}</td>"

        html += "</tr>"

    html += "</table>"
    return html


def main():
    """Main execution"""
    print("Extracting Sentence Transformers Results...")

    # Generate timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Create timestamped output directory
    output_dir = os.path.join(AGGREGATE_BASE_DIR, timestamp)
    os.makedirs(output_dir, exist_ok=True)

    print(f"Output directory: {output_dir}")

    # Extract data
    models, all_data = extract_all_data()
    print(f"Found {len(models['custom'])} custom models, {len(models['pretrained'])} pretrained models")

    # Create tables
    summary_df = create_summary_table(models, all_data)
    dataset_dfs = {}
    mteb_dfs = {}
    mteb_multilingual_dfs = {}
    mteb_multilingual_datasets_dfs = {}

    for context in CONTEXT_LENGTHS:
        dataset_dfs[context] = create_dataset_table(models, all_data, context)

        # Only create MTEB tables for 256 context
        if context == 256:
            mteb_dfs[context] = create_mteb_english_table(models, all_data, context)
            mteb_multilingual_dfs[context] = create_mteb_multilingual_table(models, all_data, context)
            mteb_multilingual_datasets_dfs[context] = create_mteb_multilingual_datasets_table(
                models, all_data, context
            )
        else:
            # Empty DataFrames for 512 context
            mteb_dfs[context] = pd.DataFrame()
            mteb_multilingual_dfs[context] = pd.DataFrame()
            mteb_multilingual_datasets_dfs[context] = pd.DataFrame()

    # Identify best performers
    summary_perf = identify_best_performers(summary_df)
    dataset_perf = {}
    for context, df in dataset_dfs.items():
        dataset_perf[f"datasets_{context}"] = identify_best_performers(df)

    # Identify best performers for MTEB tables
    mteb_perf = {}
    for context, df in mteb_dfs.items():
        if not df.empty:
            mteb_perf[f"mteb_{context}"] = identify_best_performers(df)

    mteb_multilingual_perf = {}
    for context, df in mteb_multilingual_dfs.items():
        if not df.empty:
            mteb_multilingual_perf[f"mteb_multilingual_{context}"] = identify_best_performers(df)

    mteb_multilingual_datasets_perf = {}
    for context, df in mteb_multilingual_datasets_dfs.items():
        if not df.empty:
            mteb_multilingual_datasets_perf[f"mteb_multilingual_datasets_{context}"] = identify_best_performers(df)

    perf_data = {
        "summary": summary_perf,
        **dataset_perf,
        **mteb_perf,
        **mteb_multilingual_perf,
        **mteb_multilingual_datasets_perf,
    }

    # Save files
    save_excel_with_formatting(summary_df, dataset_dfs, perf_data, output_dir)
    save_html_with_tabs(
        summary_df,
        dataset_dfs,
        mteb_dfs,
        mteb_multilingual_dfs,
        mteb_multilingual_datasets_dfs,
        perf_data,
        output_dir,
        timestamp,
    )

    # Save CSV files
    summary_df.to_csv(os.path.join(output_dir, "summary.csv"), index=False)
    for context, df in dataset_dfs.items():
        df.to_csv(os.path.join(output_dir, f"datasets_{context}.csv"), index=False)

    # Save MTEB CSV files (only for 256 context)
    if not mteb_dfs[256].empty:
        mteb_dfs[256].to_csv(os.path.join(output_dir, "mteb_256.csv"), index=False)

    if not mteb_multilingual_dfs[256].empty:
        mteb_multilingual_dfs[256].to_csv(os.path.join(output_dir, "mteb_multilingual_256.csv"), index=False)

    if not mteb_multilingual_datasets_dfs[256].empty:
        mteb_multilingual_datasets_dfs[256].to_csv(
            os.path.join(output_dir, "mteb_multilingual_datasets_256.csv"), index=False
        )

    print(f"Results saved in timestamped folder: {timestamp}")
    print("Done! 🎉")


if __name__ == "__main__":
    main()
